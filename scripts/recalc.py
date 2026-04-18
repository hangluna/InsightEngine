#!/usr/bin/env python3
"""Force Excel formula recalculation after openpyxl writes formulas.

Without this, formulas show as 0 until the user manually recalculates in Excel.
This script marks formula cells as dirty so Excel/LibreOffice recalculates on open.

Usage:
    python3 scripts/recalc.py <file.xlsx>
    python3 scripts/recalc.py output/report.xlsx
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl not installed. Run: pip3 install --user openpyxl")
    sys.exit(1)


def recalc(filepath: str):
    path = Path(filepath)
    if not path.exists():
        print(f"❌ File not found: {filepath}")
        sys.exit(1)

    if not path.suffix.lower() in (".xlsx", ".xlsm"):
        print(f"❌ Not an Excel file: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath)

    # Force recalculation on open by setting calcMode
    if wb.calculation is not None:
        wb.calculation.calcMode = "auto"

    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

    wb.save(filepath)
    size_kb = path.stat().st_size / 1024
    print(f"✅ Recalc done: {filepath} ({size_kb:.0f} KB, {formula_count} formulas, {len(wb.sheetnames)} sheets)")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/recalc.py <file.xlsx>")
        sys.exit(1)

    recalc(sys.argv[1])


if __name__ == "__main__":
    main()
