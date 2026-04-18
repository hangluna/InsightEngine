#!/usr/bin/env python3
"""Validate URLs from InsightEngine output — checks if they point to individual items
or search/listing pages.

Usage:
    python3 scripts/validate_urls.py --urls "url1" "url2" "url3"
    python3 scripts/validate_urls.py --excel output/jobs.xlsx --column "URL"
    python3 scripts/validate_urls.py --urls "url1" "url2" --json

Classification:
    ✅ DIRECT   — Individual item page (has item ID/slug)
    ❌ SEARCH   — Search results page (?q=, /search?)
    ❌ LISTING  — Platform listing page without specific item
    ❓ AMBIGUOUS — Can't determine from URL pattern alone
"""

import argparse
import json
import re
import sys
from pathlib import Path
from urllib.parse import urlparse, parse_qs

# Patterns that indicate a search/listing page (NOT a direct item page)
SEARCH_PATTERNS = [
    r"[?&]q=",
    r"[?&]query=",
    r"[?&]search=",
    r"[?&]keyword=",
    r"/search\b",
    r"/results\b",
    r"/find\b",
    r"/tim-kiem\b",
    r"/tag/",
    r"/category/",
    r"/danh-muc/",
    r"google\.\w+/search",
]

# Patterns that indicate a direct item page
DIRECT_PATTERNS = [
    r"/jobs?/\d+",           # Job by numeric ID
    r"/viec-lam/[\w-]+-\d+", # ITViec pattern
    r"/job/detail/",
    r"/product/\d+",
    r"/item/\d+",
    r"/p/[\w-]+",            # Short slug
    r"/[\w-]+-\d{4,}",      # slug-with-numeric-id
    r"/detail/\w+",
    r"/view/\w+",
]

# Generic listing paths (no specific item)
LISTING_PATTERNS = [
    r"^https?://[^/]+/jobs/?$",
    r"^https?://[^/]+/viec-lam/?$",
    r"^https?://[^/]+/products/?$",
    r"^https?://[^/]+/courses/?$",
    r"/page/\d+",
]


def classify_url(url: str) -> dict:
    """Classify a URL as DIRECT, SEARCH, LISTING, or AMBIGUOUS."""
    url_lower = url.lower().strip()

    # Check search patterns first
    for pattern in SEARCH_PATTERNS:
        if re.search(pattern, url_lower):
            return {"url": url, "classification": "SEARCH", "reason": f"Matches search pattern: {pattern}"}

    # Check listing patterns
    for pattern in LISTING_PATTERNS:
        if re.search(pattern, url_lower):
            return {"url": url, "classification": "LISTING", "reason": f"Matches listing pattern: {pattern}"}

    # Check direct patterns
    for pattern in DIRECT_PATTERNS:
        if re.search(pattern, url_lower):
            return {"url": url, "classification": "DIRECT", "reason": f"Matches direct item pattern: {pattern}"}

    # Heuristic: URL path has 3+ segments and ends with a slug → likely direct
    parsed = urlparse(url_lower)
    path_parts = [p for p in parsed.path.split("/") if p]
    if len(path_parts) >= 2 and len(path_parts[-1]) > 5:
        return {"url": url, "classification": "DIRECT", "reason": "Deep path with slug (heuristic)"}

    return {"url": url, "classification": "AMBIGUOUS", "reason": "Could not determine from URL pattern"}


def load_urls_from_excel(filepath: str, column: str) -> list[str]:
    """Extract URLs from an Excel column."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ openpyxl not installed. Run: pip3 install --user openpyxl")
        sys.exit(1)

    path = Path(filepath)
    if not path.exists():
        print(f"❌ File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find column by header
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_idx = None
    for i, header in enumerate(header_row):
        if header and column.lower() in str(header).lower():
            col_idx = i
            break

    if col_idx is None:
        print(f"❌ Column '{column}' not found. Available: {[h for h in header_row if h]}")
        sys.exit(1)

    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_idx < len(row) and row[col_idx]:
            val = str(row[col_idx]).strip()
            if val.startswith("http"):
                urls.append(val)

    return urls


def main():
    parser = argparse.ArgumentParser(description="Validate URLs from InsightEngine output")
    parser.add_argument("--urls", nargs="+", help="URLs to validate")
    parser.add_argument("--excel", help="Excel file path to extract URLs from")
    parser.add_argument("--column", default="URL", help="Column name containing URLs (default: URL)")
    parser.add_argument("--json", action="store_true", help="Output as JSON")

    args = parser.parse_args()

    if not args.urls and not args.excel:
        parser.print_help()
        sys.exit(1)

    # Collect URLs
    urls = []
    if args.urls:
        urls.extend(args.urls)
    if args.excel:
        urls.extend(load_urls_from_excel(args.excel, args.column))

    if not urls:
        print("❌ No URLs found to validate")
        sys.exit(1)

    # Classify
    results = [classify_url(url) for url in urls]

    # Count
    counts = {"DIRECT": 0, "SEARCH": 0, "LISTING": 0, "AMBIGUOUS": 0}
    for r in results:
        counts[r["classification"]] += 1

    total = len(results)
    bad = counts["SEARCH"] + counts["LISTING"]
    bad_pct = (bad / total * 100) if total > 0 else 0
    pass_threshold = bad_pct <= 30

    if args.json:
        output = {
            "total_urls": total,
            "classifications": counts,
            "bad_percentage": round(bad_pct, 1),
            "pass": pass_threshold,
            "results": results,
        }
        print(json.dumps(output, indent=2, ensure_ascii=False))
    else:
        icon_map = {"DIRECT": "✅", "SEARCH": "❌", "LISTING": "❌", "AMBIGUOUS": "❓"}
        print(f"📋 Kết quả kiểm tra {total} URLs:\n")
        for r in results:
            icon = icon_map[r["classification"]]
            print(f"  {icon} {r['classification']:10s} {r['url'][:80]}")
        print()
        print(f"  ✅ DIRECT:    {counts['DIRECT']} ({counts['DIRECT']/total*100:.0f}%)")
        print(f"  ❌ SEARCH:    {counts['SEARCH']} ({counts['SEARCH']/total*100:.0f}%)")
        print(f"  ❌ LISTING:   {counts['LISTING']} ({counts['LISTING']/total*100:.0f}%)")
        print(f"  ❓ AMBIGUOUS: {counts['AMBIGUOUS']} ({counts['AMBIGUOUS']/total*100:.0f}%)")
        print()
        if pass_threshold:
            print(f"  ✅ PASS — {bad_pct:.0f}% bad URLs (threshold: ≤30%)")
        else:
            print(f"  ❌ FAIL — {bad_pct:.0f}% bad URLs (threshold: ≤30%)")

    sys.exit(0 if pass_threshold else 1)


if __name__ == "__main__":
    main()
